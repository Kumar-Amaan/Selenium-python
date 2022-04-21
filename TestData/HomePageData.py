import openpyxl


class HomePageData:
    test_HomePage_Data = [{"name": "Aman Kumar", "email": "Aman.com", "password": "qwerty", "gender": "Male", "DOB": "06-08-1998"},
                          {"name": "Anish Kumar", "email": "Anish.com", "password": "werty", "gender": "Female", "DOB": "01-12-1999"}]

    # @staticmethod
    # def getTestData(test_case_name):
    #     Dict = {}
    #     book = openpyxl.load_workbook("C:\\Users\\FL_LPT-374\\PycharmProjects1\\SelPyFramework\\Excel\\Book1.xlsx")
    #     sheet = book.active
    #     for i in range(1, sheet.max_row + 1):  # to get rows
    #         if sheet.cell(row=i, column=1).value == test_case_name:
    #
    #             for j in range(2, sheet.max_column + 1):  # to get columns
    #                 # Dict["lastname"]="shetty
    #                 Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #     return [Dict]
